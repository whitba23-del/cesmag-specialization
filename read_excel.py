import os
import re
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR_IMPORT: openpyxl no está instalado")
    sys.exit(1)

def clean_value(val):
    if val is None:
        return ""
    s = str(val).strip()
    s = s.replace('\r\n', '\n').replace('\r', '\n')
    # Escapar comillas invertidas para no romper los template literals de JS
    s = s.replace('`', '\\`').replace('$', '\\$')
    return s

def main():
    excel_path = "docentes.xlsx"
    app_js_path = "app.js"
    
    if not os.path.exists(excel_path):
        print(f"ERROR_FILE: No se encuentra el archivo {excel_path}")
        sys.exit(2)
        
    try:
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = workbook.active
        print(f"INFO: Leyendo la hoja '{sheet.title}'")
        
        teachers = []
        
        # Iterar a partir de la fila 2
        for row_idx in range(2, sheet.max_row + 1):
            name = sheet.cell(row=row_idx, column=5).value
            
            if not name:
                continue
                
            name_str = clean_value(name)
            
            # Omitir fila de cabeceras
            if name_str.lower() in ["nombre del docente", "nombre", "docente"]:
                continue
                
            teacher = {
                "id": len(teachers) + 1,
                "name": name_str,
                "modulo": clean_value(sheet.cell(row=row_idx, column=3).value),
                "semestre": clean_value(sheet.cell(row=row_idx, column=4).value),
                "vinculacion": clean_value(sheet.cell(row=row_idx, column=6).value),
                "perfil": clean_value(sheet.cell(row=row_idx, column=2).value),
                "formacion": clean_value(sheet.cell(row=row_idx, column=7).value),
                "experiencia": clean_value(sheet.cell(row=row_idx, column=8).value),
                "pedagogia": clean_value(sheet.cell(row=row_idx, column=9).value),
                "tecnologias": clean_value(sheet.cell(row=row_idx, column=10).value),
            }
            teachers.append(teacher)
            
        print(f"SUCCESS: Se extrajeron {len(teachers)} docentes válidos del Excel.")
        
        # Generar la declaración JS en formato limpio con template literals
        js_lines = ["const TEACHERS_DATA = ["]
        for idx, t in enumerate(teachers):
            comma = "," if idx < len(teachers) - 1 else ""
            js_lines.append("  {")
            js_lines.append(f'    id: {t["id"]},')
            js_lines.append(f'    name: `{t["name"]}`,')
            js_lines.append(f'    modulo: `{t["modulo"]}`,')
            js_lines.append(f'    semestre: `{t["semestre"]}`,')
            js_lines.append(f'    vinculacion: `{t["vinculacion"]}`,')
            js_lines.append(f'    perfil: `{t["perfil"]}`,')
            js_lines.append(f'    formacion: `{t["formacion"]}`,')
            js_lines.append(f'    experiencia: `{t["experiencia"]}`,')
            js_lines.append(f'    pedagogia: `{t["pedagogia"]}`,')
            js_lines.append(f'    tecnologias: `{t["tecnologias"]}`')
            js_lines.append("  }" + comma)
        js_lines.append("];")
        
        js_replacement = "\n".join(js_lines)
        
        # Leer el archivo app.js actual
        if not os.path.exists(app_js_path):
            print(f"ERROR: No se encuentra el archivo {app_js_path}")
            sys.exit(3)
            
        with open(app_js_path, "r", encoding="utf-8") as f:
            app_js_content = f.read()
            
        # Reemplazar la constante vieja
        # Buscaremos desde "const TEACHERS_DATA = [" hasta "];" usando regex robusto
        new_content, count = re.subn(r"const TEACHERS_DATA\s*=\s*\[.*?\]\s*;\s*", js_replacement + "\n\n", app_js_content, flags=re.DOTALL)
        
        if count == 0:
            new_content, count = re.subn(r"const TEACHERS_DATA\s*=\s*\[.*?\]\s*", js_replacement + "\n", app_js_content, flags=re.DOTALL)
            
        if count > 0:
            with open(app_js_path, "w", encoding="utf-8") as f:
                f.write(new_content)
            print(f"SUCCESS: Se actualizó el archivo '{app_js_path}' exitosamente con Template Literals (seguro contra saltos de línea).")
        else:
            print("ERROR: No se pudo encontrar y reemplazar la constante en app.js")
            sys.exit(5)
            
    except Exception as e:
        print(f"ERROR: Ocurrió un error en la automatización: {str(e)}")
        sys.exit(4)

if __name__ == "__main__":
    main()
